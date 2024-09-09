from flask import Flask, render_template, request, redirect, url_for, send_file
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os

app = Flask(__name__)

# Path ke folder yang menyimpan file Excel
EXCEL_FOLDER = 'excel_files'

# Buat folder jika belum ada
os.makedirs(EXCEL_FOLDER, exist_ok=True)

def get_excel_file():
    """Ambil nama file Excel berdasarkan bulan dan tahun saat ini"""
    now = datetime.now()
    return os.path.join(EXCEL_FOLDER, f"financial_record_{now.strftime('%Y_%m')}.xlsx")

def update_saldo(pemasukan, pengeluaran):
    file_path = get_excel_file()
    
    if not os.path.exists(file_path):
        # Jika file tidak ada, buat file baru
        wb = Workbook()
        sheet = wb.active
        sheet.append(['Tanggal', 'Kategori', 'Deskripsi', 'Status', 'Pemasukan', 'Pengeluaran', 'Saldo'])
        wb.save(file_path)
    
    wb = load_workbook(file_path)
    sheet = wb.active

    # Ambil saldo terakhir dari kolom Saldo
    saldo_terakhir = sheet.cell(row=sheet.max_row, column=7).value
    saldo_terakhir = saldo_terakhir if isinstance(saldo_terakhir, (int, float)) else 0
    saldo_baru = saldo_terakhir + pemasukan - pengeluaran
    
    wb.close()
    return saldo_baru

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        tanggal = request.form['tanggal']
        kategori = request.form['kategori']
        deskripsi = request.form['deskripsi']
        status = request.form['status']
        jumlah_str = request.form['jumlah']

        try:
            jumlah = int(jumlah_str) if jumlah_str else 0
        except ValueError:
            # Jika tidak bisa mengonversi, set jumlah ke 0
            jumlah = 0

        # Tentukan apakah itu pemasukan atau pengeluaran berdasarkan status
        pemasukan = jumlah if status == "Uang masuk" else 0
        pengeluaran = jumlah if status in ["Uang keluar", "Uang disimpan"] else 0

        # Update saldo
        saldo_baru = update_saldo(pemasukan, pengeluaran)

        # Tambahkan data ke Excel
        file_path = get_excel_file()
        wb = load_workbook(file_path)
        sheet = wb.active
        sheet.append([tanggal, kategori, deskripsi, status, pemasukan, pengeluaran, saldo_baru])
        wb.save(file_path)

        return redirect(url_for('index'))

    # Ambil saldo terbaru
    file_path = get_excel_file()
    wb = load_workbook(file_path)
    sheet = wb.active
    saldo = sheet.cell(row=sheet.max_row, column=7).value
    saldo = saldo if isinstance(saldo, (int, float)) else 0
    wb.close()

    current_date = datetime.now().strftime('%Y_%m')
    return render_template('form.html', saldo=saldo, current_date=current_date)

@app.route('/download_report/<month_year>')
def download_report(month_year):
    file_path = os.path.join(EXCEL_FOLDER, f"financial_record_{month_year}.xlsx")
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return "File not found", 404

if __name__ == '__main__':
    app.run(debug=True)
